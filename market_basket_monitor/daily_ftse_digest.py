#!/usr/bin/env python3
"""Build and optionally email a property-focused daily FTSE 100 briefing."""

from __future__ import annotations

import argparse
import base64
import hashlib
import html
import json
import re
from datetime import datetime, time, timedelta, timezone
from io import StringIO
from pathlib import Path

import feedparser
import pandas as pd
import requests
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from market_monitor import fetch_news, send_email


BASE_DIR = Path(__file__).resolve().parent
CONSTITUENT_CACHE = BASE_DIR / "ftse100_constituents.json"
HTML_OUTPUT = BASE_DIR / "daily_ftse_digest.html"
WORKBOOK_OUTPUT = BASE_DIR / "daily_ftse_dashboard.xlsx"
CLOSE_CHECK_PATH = BASE_DIR / ".daily_ftse_close_check.json"
LOGO_PATH = BASE_DIR / "assets" / "melquantlab-logo.jpg"
FTSE_SOURCE = "https://en.wikipedia.org/wiki/FTSE_100_Index"
BOE_DATA_URL = "https://www.bankofengland.co.uk/boeapps/database/_iadb-fromshowcolumns.asp"
ISHARES_HOLDINGS_URL = (
    "https://www.ishares.com/uk/individual/en/products/251795/"
    "ishares-core-ftse-100-ucits-etf?siteEntryPassthrough=true&switchLocale=y"
)

CAPITAL_FINANCE = {
    "HSBA.L": "HSBC",
    "BARC.L": "Barclays",
    "LLOY.L": "Lloyds Banking Group",
    "NWG.L": "NatWest Group",
}

AUTOS_SECTOR_SUMMARY = {
    "AUTO.L": "Auto Trader",
    "INCH.L": "Inchcape",
    "AML.L": "Aston Martin Lagonda",
    "MOTR.L": "Motorpoint",
}

REAL_ESTATE_BELLWETHERS = {
    "SGRO.L": "Segro",
    "LAND.L": "Land Securities",
    "BLND.L": "British Land",
    "LMP.L": "LondonMetric Property",
    "BBOX.L": "Tritax Big Box REIT",
}

HOUSEBUILDERS = {
    "BTRW.L": "Barratt Redrow",
    "PSN.L": "Persimmon",
    "TW.L": "Taylor Wimpey",
    "BKG.L": "Berkeley Group",
    "VTY.L": "Vistry Group",
    "BWY.L": "Bellway",
}

MATERIALS_INFRASTRUCTURE = {
    "HWDN.L": "Howdens Joinery",
    "BREE.L": "Breedon Group",
    "RIO.L": "Rio Tinto",
    "GLEN.L": "Glencore",
    "AAL.L": "Anglo American",
    "CRH": "CRH (external bellwether)",
    "HG=F": "Copper futures",
}

CONSUMER_HEALTH = {
    "NXT.L": "Next",
    "MKS.L": "Marks & Spencer",
    "KGF.L": "Kingfisher",
    "TSCO.L": "Tesco",
    "SBRY.L": "Sainsbury's",
    "ULVR.L": "Unilever",
}

CONSUMER_DISCRETIONARY = {"NXT.L", "MKS.L", "KGF.L"}
CONSUMER_STAPLES = {"TSCO.L", "SBRY.L", "ULVR.L"}

SECTOR_ORDER = (
    "Financials",
    "Consumer Staples",
    "Industrials",
    "Energy",
    "Health Care",
    "Basic Materials",
    "Consumer Discretionary",
    "Utilities",
    "Technology & Media",
    "Real Estate",
)

# Current broad-sector overrides for names whose detailed ICB labels are ambiguous.
BROAD_SECTOR_OVERRIDES = {
    "III.L": "Financials",
    "MNG.L": "Financials",
    "SMT.L": "Financials",
    "PCT.L": "Financials",
    "FCIT.L": "Financials",
    "BRBY.L": "Consumer Discretionary",
    "MKS.L": "Consumer Staples",
    "REL.L": "Industrials",
    "MTLN.L": "Industrials",
    "HLMA.L": "Technology & Media",
    "SGE.L": "Technology & Media",
}

MARKET_DRIVERS = {
    "^FTSE": "FTSE 100",
    "GBPUSD=X": "GBP/USD",
    "GBPEUR=X": "GBP/EUR",
    "BZ=F": "Brent crude",
    "IGLT.L": "UK gilts ETF",
    "^VIX": "VIX",
}

PROPERTY_SECTOR_KEYWORDS = (
    "Real Estate",
    "Homebuilding",
    "Construction Supplies",
    "Construction and Materials",
)

UK_HOUSEBUILDERS = (
    "Barratt",
    "Berkeley",
    "Bellway",
    "Crest Nicholson",
    "Persimmon",
    "Redrow",
    "Taylor Wimpey",
    "Vistry",
)


def fetch_boe_property_rates() -> dict[str, dict[str, object]]:
    """Fetch the latest official SONIA and Bank Rate observations from the BoE."""
    today = datetime.now().date()
    response = requests.get(
        BOE_DATA_URL,
        params={
            "csv.x": "yes",
            "Datefrom": (today - timedelta(days=45)).strftime("%d/%b/%Y"),
            "Dateto": today.strftime("%d/%b/%Y"),
            "SeriesCodes": "IUDSOIA,IUDBEDR",
            "CSVF": "TN",
            "UsingCodes": "Y",
            "VPD": "Y",
            "VFD": "N",
        },
        timeout=20,
        headers={"User-Agent": "MelquantLabsMarketMonitor/1.0"},
    )
    response.raise_for_status()
    frame = pd.read_csv(StringIO(response.text))
    rates = {}
    for code, label in (("IUDSOIA", "SONIA"), ("IUDBEDR", "Bank Rate")):
        values = pd.to_numeric(frame.get(code), errors="coerce")
        valid = frame.loc[values.notna(), ["DATE", code]]
        if not valid.empty:
            latest = valid.iloc[-1]
            rates[label] = {
                "value": float(latest[code]),
                "as_of": str(latest["DATE"]),
            }
    return rates


def _display_sector(sector: str) -> str:
    """Combine ETF classifications into the dashboard's investment-house sectors."""
    if sector in {"Communication", "Information Technology"}:
        return "Technology & Media"
    if sector == "Materials":
        return "Basic Materials"
    return sector


def _fallback_sector(sector: str) -> str:
    """Map detailed ICB-style labels when the holdings classification is unavailable."""
    value = str(sector).lower()
    if "real estate" in value:
        return "Real Estate"
    if any(term in value for term in ("bank", "insurance", "finance", "investment")):
        return "Financials"
    if any(term in value for term in ("food", "beverage", "tobacco", "personal goods")):
        return "Consumer Staples"
    if any(term in value for term in ("oil", "gas", "coal", "energy")):
        return "Energy"
    if any(term in value for term in ("pharma", "health", "medical")):
        return "Health Care"
    if any(term in value for term in ("mining", "chemical", "industrial material")):
        return "Basic Materials"
    if any(term in value for term in ("electricity", "utilities", "water")):
        return "Utilities"
    if any(term in value for term in ("software", "computer", "technology", "media", "telecom")):
        return "Technology & Media"
    if any(term in value for term in ("retail", "travel", "leisure", "automobile", "home construction")):
        return "Consumer Discretionary"
    return "Industrials"


def fetch_ftse_sector_metadata() -> tuple[dict[str, float], dict[str, str], str]:
    """Load FTSE 100 proxy sector weights from the iShares ISF public page."""
    response = requests.get(
        ISHARES_HOLDINGS_URL,
        timeout=25,
        headers={"User-Agent": "MelquantLabsMarketMonitor/1.0"},
    )
    response.raise_for_status()
    decoded = html.unescape(response.text)
    fund_match = re.search(
        r'"fund":\{.*?"value":(\[[^\]]+\]).*?"fullName":"exposureBreakdowns\.fund"',
        decoded,
        re.DOTALL,
    )
    type_match = re.search(
        r'"type":\{.*?"value":(\[[^\]]+\]).*?"fullName":"exposureBreakdowns\.type"',
        decoded,
        re.DOTALL,
    )
    date_match = re.search(
        r'"formattedValue":"([0-9]{2}/[A-Za-z]{3}/[0-9]{4})".*?'
        r'"fullName":"exposureBreakdowns\.asOf"',
        decoded,
        re.DOTALL,
    )
    if not fund_match or not type_match:
        raise ValueError("iShares sector breakdown was not found")
    values = json.loads(fund_match.group(1))
    sectors = json.loads(type_match.group(1))
    weights = {}
    for sector, value in zip(sectors, values):
        if sector == "Cash and/or Derivatives":
            continue
        display_sector = _display_sector(sector)
        weights[display_sector] = weights.get(display_sector, 0.0) + float(value)
    return weights, {}, date_match.group(1) if date_match else "latest available"


def property_exposure(frame: pd.DataFrame) -> pd.DataFrame:
    """Select FTSE companies most relevant to property development and construction."""
    if frame.empty or "sector" not in frame.columns:
        return frame.iloc[0:0].copy()
    sector_pattern = "|".join(PROPERTY_SECTOR_KEYWORDS)
    sector_match = frame["sector"].str.contains(sector_pattern, case=False, na=False)
    company_match = frame["company"].str.contains("|".join(UK_HOUSEBUILDERS), case=False, na=False)
    return frame[sector_match | company_match].copy()


def yahoo_lse_ticker(epic: str) -> str:
    """Convert a London Stock Exchange EPIC into Yahoo Finance format."""
    symbol = str(epic).strip().upper()
    if symbol.endswith("."):
        symbol = symbol[:-1]
    symbol = symbol.replace(".", "-")
    return f"{symbol}.L"


def _flatten_column(column) -> str:
    if isinstance(column, tuple):
        return " ".join(str(part) for part in column if str(part) != "nan").strip()
    return str(column).strip()


def load_ftse100_constituents() -> list[dict[str, str]]:
    """Load the current FTSE 100 table and retain a local fallback cache."""
    try:
        response = requests.get(
            FTSE_SOURCE,
            timeout=20,
            headers={"User-Agent": "MelquantLabsMarketMonitor/1.0"},
        )
        response.raise_for_status()
        tables = pd.read_html(StringIO(response.text))
        selected = None
        for table in tables:
            table = table.copy()
            table.columns = [_flatten_column(column) for column in table.columns]
            ticker_column = next(
                (column for column in table.columns if column.lower() in {"ticker", "epic"}),
                None,
            )
            company_column = next(
                (column for column in table.columns if column.lower() == "company"),
                None,
            )
            if ticker_column and company_column and len(table) >= 90:
                selected = (table, ticker_column, company_column)
                break
        if selected is None:
            raise ValueError("FTSE 100 constituent table was not found")

        table, ticker_column, company_column = selected
        sector_column = next(
            (column for column in table.columns if "sector" in column.lower()),
            None,
        )
        constituents = []
        for _, row in table.iterrows():
            epic = str(row[ticker_column]).strip()
            company = str(row[company_column]).strip()
            if not epic or epic.lower() == "nan":
                continue
            constituents.append(
                {
                    "ticker": yahoo_lse_ticker(epic),
                    "epic": epic,
                    "company": company,
                    "sector": (
                        str(row[sector_column]).strip()
                        if sector_column and pd.notna(row[sector_column])
                        else "Unclassified"
                    ),
                }
            )
        if len(constituents) < 90:
            raise ValueError("FTSE 100 source returned too few constituents")
        CONSTITUENT_CACHE.write_text(json.dumps(constituents, indent=2), encoding="utf-8")
        return constituents
    except Exception:
        if CONSTITUENT_CACHE.exists():
            return json.loads(CONSTITUENT_CACHE.read_text(encoding="utf-8"))
        raise


def fetch_snapshot(tickers: list[str]) -> pd.DataFrame:
    """Download a batch and calculate daily, weekly and volume measures."""
    unique_tickers = list(dict.fromkeys(tickers))
    data = yf.download(
        unique_tickers,
        period="2mo",
        interval="1d",
        auto_adjust=True,
        actions=False,
        progress=False,
        group_by="ticker",
        threads=True,
    )
    rows = []
    for ticker in unique_tickers:
        try:
            frame = data[ticker] if isinstance(data.columns, pd.MultiIndex) else data
            close = pd.to_numeric(frame["Close"], errors="coerce").dropna()
            if len(close) < 2:
                continue
            volume = pd.to_numeric(frame.get("Volume"), errors="coerce").dropna()
            average_volume = float(volume.tail(20).mean()) if not volume.empty else 0.0
            latest_volume = float(volume.iloc[-1]) if not volume.empty else 0.0
            rows.append(
                {
                    "ticker": ticker,
                    "price": float(close.iloc[-1]),
                    "day_pct": float((close.iloc[-1] / close.iloc[-2] - 1) * 100),
                    "week_pct": (
                        float((close.iloc[-1] / close.iloc[-6] - 1) * 100)
                        if len(close) >= 6
                        else None
                    ),
                    "month_pct": (
                        float((close.iloc[-1] / close.iloc[-22] - 1) * 100)
                        if len(close) >= 22
                        else None
                    ),
                    "volume_ratio": (
                        latest_volume / average_volume if average_volume > 0 else None
                    ),
                    "as_of": close.index[-1].date().isoformat(),
                }
            )
        except (KeyError, TypeError, ValueError, IndexError):
            continue
    return pd.DataFrame(rows).set_index("ticker") if rows else pd.DataFrame()


def attach_names(snapshot: pd.DataFrame, names: dict[str, str]) -> pd.DataFrame:
    result = snapshot.copy()
    result["company"] = [names.get(ticker, ticker) for ticker in result.index]
    return result


def notable_moves(snapshot: pd.DataFrame) -> pd.DataFrame:
    if snapshot.empty:
        return snapshot
    volume = snapshot["volume_ratio"].fillna(0)
    return snapshot[(snapshot["day_pct"].abs() >= 3.0) | (volume >= 2.0)].copy()


TRUSTED_NEWS_SOURCES = {
    "bank of england": 100,
    "office for national statistics": 100,
    "reuters": 100,
    "financial times": 95,
    "bloomberg": 95,
    "gov.uk": 95,
    "london stock exchange": 90,
    "rics": 90,
    "yahoo finance": 90,
    "bbc": 85,
    "sky news": 80,
    "the guardian": 75,
    "city a.m.": 75,
    "property week": 75,
    "the construction index": 75,
    "this is money": 70,
    "proactive investors": 65,
}

LOW_VALUE_HEADLINE_TERMS = (
    "i'm buying",
    "i’m buying",
    "juicy dividends",
    "should you buy",
    "could make you rich",
    "passive income stock",
)


def _normalise_story(story: dict) -> dict:
    """Separate Google News' source suffix and preserve source attribution."""
    result = dict(story)
    title = str(result.get("title", "")).strip()
    source = str(result.get("source", "") or "").strip()
    if not source and " - " in title:
        title, source = title.rsplit(" - ", 1)
    elif source and title.endswith(f" - {source}"):
        title = title[: -(len(source) + 3)]
    result["title"] = title.strip()
    result["source"] = source.strip()
    return result


def _story_rank(story: dict) -> tuple[int, float]:
    source = str(story.get("source", "")).lower()
    trust_score = max(
        (score for name, score in TRUSTED_NEWS_SOURCES.items() if name in source),
        default=50,
    )
    published = story.get("published")
    timestamp = published.timestamp() if published else 0.0
    return trust_score, timestamp


def collect_ranked_news(company_names: list[str], limit: int = 3) -> list[dict]:
    """Collect and rank recent headlines from reputable publishers."""
    queries = [
        "FTSE 100 earnings CEO trading update",
        "UK property housebuilders mortgage rates construction Reuters Bloomberg Financial Times",
        "UK banks lending housing market materials prices",
    ]
    queries.extend(company_names[:5])
    stories = []
    seen = set()
    for query in queries:
        for raw_story in fetch_news(query, max_items=5, days_back=2):
            story = _normalise_story(raw_story)
            key = story["title"].lower().strip()
            if key in seen:
                continue
            seen.add(key)
            stories.append(story)
    reputable = [
        story
        for story in stories
        if _story_rank(story)[0] >= 70
        and not any(term in story["title"].lower() for term in LOW_VALUE_HEADLINE_TERMS)
    ]
    return sorted(reputable, key=_story_rank, reverse=True)[:limit]


def collect_forward_watch(limit: int = 3) -> list[dict]:
    """Find recent forward-looking headlines without inventing calendar events."""
    forward_terms = (
        "week ahead",
        "what to watch",
        "set to report",
        "due to report",
        "earnings calendar",
        "results calendar",
        "upcoming",
        "next week",
        "tomorrow",
        "outlook",
        "forecast",
        "expected to",
    )
    relevance_terms = (
        "ftse",
        "uk ",
        "britain",
        "british",
        "london",
        "sterling",
        "property",
        "housing",
        "housebuilder",
        "mortgage",
        "bank of england",
        "construction",
        "planning",
        "rics",
    )
    queries = [
        "FTSE 100 earnings this week outlook Reuters OR Bloomberg OR Financial Times",
        "UK market week ahead earnings economic data",
        "UK property housing construction week ahead rates planning RICS",
    ]
    stories = []
    seen = set()
    for query in queries:
        for raw_story in fetch_news(query, max_items=5, days_back=7):
            story = _normalise_story(raw_story)
            key = story["title"].lower().strip()
            if (
                not key
                or key in seen
                or not any(term in key for term in forward_terms)
                or not any(term in key for term in relevance_terms)
            ):
                continue
            seen.add(key)
            stories.append(story)
    return sorted(stories, key=_story_rank, reverse=True)[:limit]


def build_pm_summary(
    ftse: pd.DataFrame,
    sector_moves: pd.DataFrame,
    drivers: pd.DataFrame,
    banks: pd.DataFrame,
    autos: pd.DataFrame,
    technology: pd.DataFrame,
    materials: pd.DataFrame,
    real_estate: pd.DataFrame,
    housebuilders: pd.DataFrame,
    consumers: pd.DataFrame,
    property_rates: dict[str, dict[str, object]],
    news: list[dict],
    forward_news: list[dict],
) -> list[tuple[str, str]]:
    """Create a property-developer closing note capped safely below 500 words."""
    ftse_move = drivers.loc["^FTSE", "day_pct"] if "^FTSE" in drivers.index else None
    advancers = int((ftse["day_pct"] > 0).sum()) if not ftse.empty else 0
    decliners = int((ftse["day_pct"] < 0).sum()) if not ftse.empty else 0
    gainers = ftse.sort_values("day_pct", ascending=False) if not ftse.empty else ftse
    fallers = ftse.sort_values("day_pct") if not ftse.empty else ftse
    leader = str(gainers.iloc[0]["company"]) if not gainers.empty else "the leading constituent"
    leader_move = gainers.iloc[0]["day_pct"] if not gainers.empty else None
    laggard = str(fallers.iloc[0]["company"]) if not fallers.empty else "the weakest constituent"
    laggard_move = fallers.iloc[0]["day_pct"] if not fallers.empty else None
    market_tone = "advanced" if ftse_move is not None and ftse_move > 0 else "declined" if ftse_move is not None and ftse_move < 0 else "finished broadly unchanged"

    sections = [
        (
            "Today's Market",
            f"The FTSE 100 {market_tone} {_format_pct(ftse_move)}, with {advancers} advancers "
            f"against {decliners} decliners. {leader} led the index at {_format_pct(leader_move)}, "
            f"while {laggard} was the weakest name at {_format_pct(laggard_move)}."
        )
    ]

    autos_move = autos["day_pct"].mean() if not autos.empty else None
    if not autos.empty:
        autos_leader = autos.sort_values("day_pct", ascending=False).iloc[0]
        autos_detail = (
            f" The strongest tracked name was {autos_leader['company']} at "
            f"{_format_pct(autos_leader['day_pct'])}."
        )
    else:
        autos_detail = ""
    sections.append(
        (
            "Autos",
            f"The tracked UK autos and marketplace basket averaged {_format_pct(autos_move)}."
            f"{autos_detail}",
        )
    )

    bank_move = banks["day_pct"].mean() if not banks.empty else None
    sections.append(
        (
            "Financials",
            f"The tracked UK lender basket averaged {_format_pct(bank_move)}. Its direction is a "
            "daily signal for credit sentiment, but lending standards and mortgage pricing remain "
            "the more direct property indicators.",
        )
    )

    technology_move = technology["day_pct"].mean() if not technology.empty else None
    sections.append(
        (
            "Technology",
            f"FTSE technology-related constituents averaged {_format_pct(technology_move)}. The "
            "sector is a smaller part of the UK index, so treat it mainly as a risk-appetite and "
            "business-investment signal.",
        )
    )

    gilt_move = drivers.loc["IGLT.L", "day_pct"] if "IGLT.L" in drivers.index else None
    sterling_move = drivers.loc["GBPUSD=X", "day_pct"] if "GBPUSD=X" in drivers.index else None
    brent_move = drivers.loc["BZ=F", "day_pct"] if "BZ=F" in drivers.index else None
    sections.append(
        (
            "Macro View",
            f"The FTSE 100 moved {_format_pct(ftse_move)}, the UK gilt proxy {_format_pct(gilt_move)}, "
            f"sterling {_format_pct(sterling_move)} against the dollar and Brent {_format_pct(brent_move)}. "
            "Together they frame risk appetite, financing conditions and imported-cost pressure.",
        )
    )

    if not sector_moves.empty:
        strongest_sector = str(sector_moves.index[0])
        weakest_sector = str(sector_moves.index[-1])
        sections.append(
            (
                "Sector Read",
                f"{strongest_sector} showed the strongest average performance, while "
                f"{weakest_sector} lagged. Compare this breadth with the individual stock moves "
                "above before treating an outsized move as company-specific.",
            )
        )

    sonia = property_rates.get("SONIA", {}).get("value")
    bank_rate = property_rates.get("Bank Rate", {}).get("value")
    sections.append(
        (
            "Capital & Finance",
            f"SONIA is {'unavailable' if sonia is None else f'{sonia:.4f}%'} and Bank Rate is "
            f"{'unavailable' if bank_rate is None else f'{bank_rate:.2f}%'}. The UK gilt proxy "
            f"moved {_format_pct(gilt_move)}, while the tracked lender basket averaged "
            f"{_format_pct(bank_move)}. Read these together for refinancing pressure, credit "
            "availability and buyer mortgage affordability.",
        )
    )

    materials_move = materials["day_pct"].mean() if not materials.empty else None
    copper_move = materials.loc["HG=F", "day_pct"] if "HG=F" in materials.index else None
    sections.append(
        (
            "Build-Cost Pipeline",
            f"The materials and infrastructure basket averaged {_format_pct(materials_move)}; "
            f"copper moved {_format_pct(copper_move)}. Mining shares are macro proxies, so "
            "confirm actual tender prices, labour costs and supplier quotes before changing a "
            "development budget.",
        )
    )

    reit_move = real_estate["day_pct"].mean() if not real_estate.empty else None
    builder_move = housebuilders["day_pct"].mean() if not housebuilders.empty else None
    discretionary = consumers.loc[consumers.index.intersection(CONSUMER_DISCRETIONARY)]
    staples = consumers.loc[consumers.index.intersection(CONSUMER_STAPLES)]
    discretionary_move = discretionary["day_pct"].mean() if not discretionary.empty else None
    staples_move = staples["day_pct"].mean() if not staples.empty else None
    sections.append(
        (
            "Demand & Competition",
            f"Listed real-estate bellwethers averaged {_format_pct(reit_move)} and housebuilders "
            f"averaged {_format_pct(builder_move)}. Consumer discretionary proxies averaged "
            f"{_format_pct(discretionary_move)} versus {_format_pct(staples_move)} for staples—"
            "a useful daily sentiment check, not a substitute for mortgage approvals, wages or "
            "local transaction data.",
        )
    )

    if news:
        headline_text = "; ".join(
            f"{story['title']} ({story.get('source') or 'source linked above'})"
            for story in news[:3]
        )
        sections.append(
            ("What to Pay Attention To", f"The principal reported catalysts were: {headline_text}.")
        )

    watch_items = []
    if not sector_moves.empty:
        watch_items.append(f"weakness in {sector_moves.index[-1]}")
    if not fallers.empty:
        watch_items.append(f"whether {laggard} stabilises after {_format_pct(laggard_move)}")
    if "GBPUSD=X" in drivers.index:
        watch_items.append(
            f"sterling after a {_format_pct(drivers.loc['GBPUSD=X', 'day_pct'])} move against the dollar"
        )
    if not housebuilders.empty:
        weakest_builder = housebuilders.sort_values("day_pct").iloc[0]
        watch_items.append(
            f"{weakest_builder['company']} after {_format_pct(weakest_builder['day_pct'])}"
        )
    watch_text = "; ".join(watch_items[:4])
    if watch_text:
        watch_text = watch_text[0].upper() + watch_text[1:] + "."
    else:
        watch_text = "No material cross-market warning signal was identified in today's data."
    sections.append(("Things We're Watching", watch_text))

    if forward_news:
        watch_text = "; ".join(
            f"{story['title']} ({story.get('source') or 'source linked below'})"
            for story in forward_news[:3]
        )
        sections.append(
            (
                "Prepare for the Next Session",
                f"Monitor: {watch_text}. Treat these as a watchlist rather than a confirmed "
                "event calendar, and verify timing at the linked source.",
            )
        )
    else:
        sections.append(
            (
                "Prepare for the Next Session",
                "Check scheduled company statements, Bank of England communications and UK macro "
                "releases before the open. Review mortgage approvals, CPI, wages, planning data "
                "and RICS releases when due; confirm dates with the original publisher.",
            )
        )

    priority_titles = (
        "What to Pay Attention To",
        "Things We're Watching",
        "Prepare for the Next Session",
    )
    priority = [section for title in priority_titles for section in sections if section[0] == title]
    remaining = [section for section in sections if section[0] not in priority_titles]
    sections = priority + remaining

    words = " ".join(body for _, body in sections).split()
    if len(words) > 480:
        return [("Daily Briefing", " ".join(words[:480]).rstrip(".,;:") + ".")]
    return sections


def _format_pct(value) -> str:
    return "—" if pd.isna(value) else f"{value:+.2f}%"


def _move_colour(value) -> str:
    if pd.isna(value):
        return "#9FB0C5"
    if value > 0:
        return "#34D6A2"
    if value < 0:
        return "#FF6B7A"
    return "#F7FAFC"


def _compact_text(value: str, limit: int = 165) -> str:
    """Keep executive-strip copy concise without cutting through a word."""
    cleaned = " ".join(value.split())
    if len(cleaned) <= limit:
        return cleaned
    shortened = cleaned[: limit - 1].rsplit(" ", 1)[0].rstrip(".,;:")
    return shortened + "…"


def _table(
    frame: pd.DataFrame,
    columns: list[tuple[str, str]],
    limit=None,
    emphasise_first: bool = False,
    emphasise_extremes: bool = False,
) -> str:
    if frame.empty:
        return '<p style="color:#9FB0C5;margin:12px 0 22px;">No usable market data was returned.</p>'
    shown = frame.head(limit) if limit else frame
    header = "".join(
        '<th style="background:#17395F;color:#D9E6F3;font-size:11px;letter-spacing:.5px;'
        f'text-transform:uppercase;padding:10px 9px;text-align:left;border-bottom:1px solid #31567E;">{html.escape(label)}</th>'
        for _, label in columns
    )
    rows = []
    for row_number, (ticker, row) in enumerate(shown.iterrows()):
        emphasised = (emphasise_first and row_number == 0) or (
            emphasise_extremes and row_number in {0, len(shown) - 1}
        )
        cells = []
        for column, _ in columns:
            cell_colour = "#ECF4FC"
            weight = "800" if emphasised else "500"
            if column == "ticker":
                value = ticker
            elif column in {"day_pct", "week_pct", "month_pct"}:
                raw_value = row.get(column)
                value = _format_pct(raw_value)
                cell_colour = _move_colour(raw_value)
                weight = "800" if emphasised else "700"
            elif column == "volume_ratio":
                ratio = row.get(column)
                value = "—" if pd.isna(ratio) else f"{ratio:.1f}x"
                if not pd.isna(ratio) and ratio >= 2:
                    cell_colour = "#F6BD4A"
                    weight = "800" if emphasised else "700"
            elif column == "price":
                value = f"{row.get(column):.2f}"
            else:
                value = str(row.get(column, ""))
            background = "#102A47" if row_number % 2 == 0 else "#0D243E"
            cells.append(
                f'<td style="background:{background};color:{cell_colour};font-size:13px;'
                f'font-weight:{weight};padding:10px 9px;border-bottom:1px solid #274766;">'
                f'{html.escape(value)}</td>'
            )
        rows.append("<tr>" + "".join(cells) + "</tr>")
    return (
        '<table class="data-table" role="presentation" cellspacing="0" cellpadding="0" width="100%" '
        'style="width:100%;table-layout:fixed;border-collapse:separate;border-spacing:0;margin:10px 0 24px;border:1px solid #274766;'
        f'border-radius:10px;overflow:hidden;"><thead><tr>{header}</tr></thead><tbody>{"".join(rows)}</tbody></table>'
    )


def _metric_card(label: str, value: str, colour: str, note: str = "") -> str:
    return f"""
    <td class="metric-cell" width="25%" valign="top" style="padding:5px;">
      <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#112C4A;border:1px solid #2B4F73;border-radius:12px;">
        <tr><td style="padding:14px 14px 4px;color:#8EA5BD;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;">{html.escape(label)}</td></tr>
        <tr><td style="padding:0 14px;color:{colour};font-size:24px;font-weight:700;line-height:1.2;">{html.escape(value)}</td></tr>
        <tr><td style="padding:5px 14px 14px;color:#8EA5BD;font-size:10px;">{html.escape(note)}</td></tr>
      </table>
    </td>"""


def _section_title(kicker: str, title: str) -> str:
    return f"""
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin-top:28px;">
      <tr><td style="color:#F6BD4A;font-size:10px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;padding-bottom:5px;">{html.escape(kicker)}</td></tr>
      <tr><td style="color:#F7FAFC;font-size:23px;font-weight:700;line-height:1.2;padding-bottom:9px;border-bottom:1px solid #345675;">{html.escape(title)}</td></tr>
    </table>"""


def build_alpha_view(
    ftse: pd.DataFrame,
    broad_sector_frames: dict[str, pd.DataFrame],
    real_estate: pd.DataFrame,
    housebuilders: pd.DataFrame,
    property_rates: dict[str, dict[str, object]],
    news: list[dict],
    forward_news: list[dict],
) -> list[dict[str, str]]:
    """Create three time-horizon research hypotheses from the latest tape and catalysts."""
    ranked = ftse.sort_values("day_pct", ascending=False) if not ftse.empty else ftse
    leader = ranked.iloc[0] if not ranked.empty else None
    headline = news[0]["title"] if news else "No qualifying company catalyst was returned"
    source = news[0].get("source") if news else "market tape"
    company_tape = pd.concat([ftse, real_estate, housebuilders]).drop_duplicates()
    headline_lower = headline.lower()
    matched = company_tape[
        company_tape["company"].astype(str).map(lambda name: name.lower() in headline_lower)
    ] if not company_tape.empty else company_tape
    day_name = matched.iloc[0] if not matched.empty else leader
    day_subject = str(day_name["company"]) if day_name is not None else "FTSE leadership"
    day_move = _format_pct(day_name["day_pct"]) if day_name is not None else "—"

    sector_averages = {
        sector: frame["day_pct"].mean()
        for sector, frame in broad_sector_frames.items()
        if not frame.empty
    }
    strongest = max(sector_averages, key=sector_averages.get) if sector_averages else "FTSE leaders"
    weakest = min(sector_averages, key=sector_averages.get) if sector_averages else "FTSE laggards"
    next_catalyst = forward_news[0]["title"] if forward_news else "the next UK macro and company calendar"
    fortnight_catalysts = "; ".join(story["title"] for story in forward_news[:3])

    property_tape = pd.concat([real_estate, housebuilders]).drop_duplicates()
    property_tape = property_tape.sort_values("month_pct", ascending=False) if not property_tape.empty else property_tape
    property_leader = property_tape.iloc[0] if not property_tape.empty else None
    property_name = str(property_leader["company"]) if property_leader is not None else "listed property and housebuilders"
    property_month = _format_pct(property_leader["month_pct"]) if property_leader is not None else "—"
    sonia = property_rates.get("SONIA", {}).get("value")
    sonia_text = "unavailable" if sonia is None else f"{sonia:.4f}%"

    return [
        {
            "horizon": "1 Day",
            "title": f"Catalyst follow-through: {day_subject}",
            "thesis": f"{headline} ({source or 'linked publisher'}). {day_subject} moved {day_move}; watch whether volume and breadth confirm the move at the next open.",
            "risk": "The catalyst may already be priced, the headline timing may be stale, or a broader FTSE reversal may overwhelm the company signal.",
        },
        {
            "horizon": "1 Week",
            "title": f"Sector rotation: {strongest} versus {weakest}",
            "thesis": f"{strongest} led the daily sector tape while {weakest} lagged. Test whether that spread persists or mean-reverts around {next_catalyst}.",
            "risk": "One session does not establish a trend; macro data, earnings and index concentration can quickly reverse the relationship.",
        },
        {
            "horizon": "14 Days",
            "title": "Verified catalyst window" if fortnight_catalysts else "No verified 14-day catalyst",
            "thesis": (
                f"Monitor these attributed forward-looking reports: {fortnight_catalysts}. Confirm each date at the linked publisher before acting."
                if fortnight_catalysts
                else "The free feeds returned no specific, attributable two-week event. No investment hypothesis is asserted for this horizon."
            ),
            "risk": "Headline feeds are not an official earnings calendar; event dates can change and must be checked against company RNS or an official release calendar.",
        },
        {
            "horizon": "1 Month",
            "title": f"Rates and property repricing: {property_name}",
            "thesis": f"{property_name} has returned {property_month} over one month while SONIA stands at {sonia_text}. Track whether easing finance conditions support property and housebuilder relative performance.",
            "risk": "Mortgage demand, planning data, refinancing costs and company statements may diverge from the listed-equity signal.",
        },
    ]


def _alpha_view_html(ideas: list[dict[str, str]]) -> str:
    colours = {
        "1 Day": "#F6BD4A",
        "1 Week": "#46CFF5",
        "14 Days": "#34D6A2",
        "1 Month": "#A78BFA",
    }
    cards = []
    for idea in ideas:
        colour = colours.get(idea["horizon"], "#F6BD4A")
        cards.append(
            '<table role="presentation" width="100%" cellspacing="0" cellpadding="0" '
            'style="background:#102A47;border:1px solid #31567E;border-radius:12px;margin:10px 0;">'
            '<tr><td style="padding:16px 18px;">'
            f'<div style="color:{colour};font-size:10px;font-weight:800;letter-spacing:1.3px;text-transform:uppercase;">{html.escape(idea["horizon"])}</div>'
            f'<div style="color:#F7FAFC;font-size:17px;font-weight:800;margin-top:5px;">{html.escape(idea["title"])}</div>'
            f'<div style="color:#DCE8F4;font-size:13px;line-height:1.55;margin-top:8px;"><strong>Hypothesis:</strong> {html.escape(idea["thesis"])}</div>'
            f'<div style="color:#9FB0C5;font-size:12px;line-height:1.5;margin-top:7px;"><strong style="color:#FF9AA5;">Key risk:</strong> {html.escape(idea["risk"])}</div>'
            '</td></tr></table>'
        )
    return "".join(cards)


def build_digest() -> tuple[str, dict[str, pd.DataFrame], list[dict]]:
    constituents = load_ftse100_constituents()
    ftse_names = {item["ticker"]: item["company"] for item in constituents}
    sectors = {item["ticker"]: item["sector"] for item in constituents}

    all_tickers = (
        list(ftse_names)
        + list(CAPITAL_FINANCE)
        + list(AUTOS_SECTOR_SUMMARY)
        + list(REAL_ESTATE_BELLWETHERS)
        + list(HOUSEBUILDERS)
        + list(MATERIALS_INFRASTRUCTURE)
        + list(CONSUMER_HEALTH)
        + list(MARKET_DRIVERS)
    )
    snapshot = fetch_snapshot(all_tickers)
    ftse = attach_names(snapshot.loc[snapshot.index.intersection(ftse_names)], ftse_names)
    ftse["sector"] = [sectors.get(ticker, "Unclassified") for ticker in ftse.index]
    banks = attach_names(
        snapshot.loc[snapshot.index.intersection(CAPITAL_FINANCE)], CAPITAL_FINANCE
    )
    autos = attach_names(
        snapshot.loc[snapshot.index.intersection(AUTOS_SECTOR_SUMMARY)],
        AUTOS_SECTOR_SUMMARY,
    )
    real_estate = attach_names(
        snapshot.loc[snapshot.index.intersection(REAL_ESTATE_BELLWETHERS)],
        REAL_ESTATE_BELLWETHERS,
    )
    housebuilders = attach_names(
        snapshot.loc[snapshot.index.intersection(HOUSEBUILDERS)], HOUSEBUILDERS
    )
    materials = attach_names(
        snapshot.loc[snapshot.index.intersection(MATERIALS_INFRASTRUCTURE)],
        MATERIALS_INFRASTRUCTURE,
    )
    consumers = attach_names(
        snapshot.loc[snapshot.index.intersection(CONSUMER_HEALTH)], CONSUMER_HEALTH
    )
    drivers = attach_names(
        snapshot.loc[snapshot.index.intersection(MARKET_DRIVERS)], MARKET_DRIVERS
    )
    technology = ftse[
        ftse["sector"].str.contains("technology|software|computer|electronic", case=False, na=False)
    ].copy()
    try:
        property_rates = fetch_boe_property_rates()
    except Exception:
        property_rates = {}
    try:
        sector_weights, ticker_sectors, sector_weights_as_of = fetch_ftse_sector_metadata()
    except Exception:
        sector_weights, ticker_sectors, sector_weights_as_of = {}, {}, "unavailable"
    ftse["broad_sector"] = [
        ticker_sectors.get(
            ticker,
            BROAD_SECTOR_OVERRIDES.get(ticker, _fallback_sector(sectors.get(ticker, ""))),
        )
        for ticker in ftse.index
    ]
    broad_sector_frames = {
        sector: ftse[ftse["broad_sector"].eq(sector)].copy()
        for sector in SECTOR_ORDER
    }

    gainers = ftse.sort_values("day_pct", ascending=False).head(5)
    fallers = ftse.sort_values("day_pct").head(5)
    notable = notable_moves(ftse).sort_values("day_pct", ascending=False)
    sector_moves = (
        ftse.groupby("sector", dropna=False)["day_pct"]
        .mean()
        .sort_values(ascending=False)
        .to_frame("day_pct")
    )
    property_names = property_exposure(ftse).sort_values("day_pct", ascending=False)
    news_names = (
        real_estate["company"].tolist()
        + housebuilders["company"].tolist()
        + (notable["company"].tolist() if not notable.empty else gainers["company"].tolist())
    )
    news = collect_ranked_news(news_names)
    forward_news = collect_forward_watch()
    ftse_index_move = drivers.loc["^FTSE", "day_pct"] if "^FTSE" in drivers.index else None

    news_html = "".join(
        f"""
        <tr>
          <td width="42" valign="top" style="padding:12px 0;">
            <div style="width:30px;height:30px;line-height:30px;text-align:center;background:#F6BD4A;color:#07121F;border-radius:15px;font-size:12px;font-weight:800;">{number:02d}</div>
          </td>
          <td valign="top" style="padding:12px 0;border-bottom:1px solid #294B6B;">
            <div style="color:#F7FAFC;font-size:15px;font-weight:700;line-height:1.35;">{html.escape(story["title"])}</div>
            <div style="color:#F6BD4A;font-size:10px;font-weight:800;letter-spacing:.7px;text-transform:uppercase;margin-top:5px;">{html.escape(story.get("source") or "Linked publisher")}</div>
            <div style="color:#9FB0C5;font-size:12px;line-height:1.45;margin-top:5px;">Potential FTSE, property, financing, materials or demand catalyst. Verify the detail and timing at source.</div>
            <a href="{html.escape(story["link"], quote=True)}" style="display:inline-block;color:#46CFF5;font-size:12px;font-weight:700;text-decoration:none;margin-top:7px;">READ SOURCE →</a>
          </td>
        </tr>"""
        for number, story in enumerate(news, start=1)
    ) or '<tr><td style="color:#9FB0C5;padding:16px 0;">No qualifying recent stories were returned.</td></tr>'

    columns = [
        ("company", "Company"),
        ("ticker", "Ticker"),
        ("day_pct", "1 day"),
        ("week_pct", "5 days"),
        ("volume_ratio", "Volume / 20d"),
    ]
    generated_at = datetime.now().astimezone()
    generated = generated_at.strftime("%A %d %B %Y, %H:%M %Z")
    report_title = (
        "Daily Market Briefing"
        if generated_at.time().replace(tzinfo=None) >= time(16, 50)
        else "INTRADAY PREVIEW — NOT MARKET CLOSE"
    )
    market_dates = ftse["as_of"].dropna().astype(str) if "as_of" in ftse else pd.Series(dtype=str)
    market_as_of = market_dates.mode().iloc[0] if not market_dates.empty else "unavailable"
    news_heading = "Three stories that matter" if len(news) == 3 else "Stories that matter"
    ftse_move = ftse_index_move
    advancers = int((ftse["day_pct"] > 0).sum())
    decliners = int((ftse["day_pct"] < 0).sum())
    leader_name = str(gainers.iloc[0]["company"]) if not gainers.empty else "—"
    leader_move = gainers.iloc[0]["day_pct"] if not gainers.empty else None
    laggard_name = str(fallers.iloc[0]["company"]) if not fallers.empty else "—"
    laggard_move = fallers.iloc[0]["day_pct"] if not fallers.empty else None
    sonia = property_rates.get("SONIA", {}).get("value")
    sonia_as_of = property_rates.get("SONIA", {}).get("as_of", "official latest")
    bank_rate = property_rates.get("Bank Rate", {}).get("value")
    bank_rate_as_of = property_rates.get("Bank Rate", {}).get("as_of", "official latest")
    gilt_move = drivers.loc["IGLT.L", "day_pct"] if "IGLT.L" in drivers.index else None
    sterling_move = drivers.loc["GBPUSD=X", "day_pct"] if "GBPUSD=X" in drivers.index else None
    bank_average = banks["day_pct"].mean() if not banks.empty else None
    materials_average = materials["day_pct"].mean() if not materials.empty else None
    real_estate_average = real_estate["day_pct"].mean() if not real_estate.empty else None
    housebuilder_average = housebuilders["day_pct"].mean() if not housebuilders.empty else None
    discretionary = consumers.loc[consumers.index.intersection(CONSUMER_DISCRETIONARY)]
    staples = consumers.loc[consumers.index.intersection(CONSUMER_STAPLES)]
    discretionary_average = discretionary["day_pct"].mean() if not discretionary.empty else None
    staples_average = staples["day_pct"].mean() if not staples.empty else None
    property_read = (
        f"FTSE real-estate bellwethers averaged {_format_pct(real_estate_average)} and the housebuilder "
        f"basket averaged {_format_pct(housebuilder_average)}. Lenders averaged {_format_pct(bank_average)} "
        f"and materials proxies averaged {_format_pct(materials_average)}. For Lotus Noor Developments, read financing, "
        "build costs and buyer demand together—no single share-price move is a development decision."
    )
    sector_display = pd.concat([sector_moves.head(4), sector_moves.tail(4)]).drop_duplicates()
    alpha_view_html = _alpha_view_html(
        build_alpha_view(
            ftse,
            broad_sector_frames,
            real_estate,
            housebuilders,
            property_rates,
            news,
            forward_news,
        )
    )
    weight_strip = " · ".join(
        f"{sector} {weight:.1f}%"
        for sector, weight in sorted(sector_weights.items(), key=lambda item: item[1], reverse=True)
    ) or "Sector weights unavailable"
    pm_summary = build_pm_summary(
        ftse,
        sector_moves,
        drivers,
        banks,
        autos,
        technology,
        materials,
        real_estate,
        housebuilders,
        consumers,
        property_rates,
        news,
        forward_news,
    )
    pm_sections = dict(pm_summary)
    executive_items = (
        ("What matters", _compact_text(pm_sections.get("What to Pay Attention To", "Review the principal FTSE and property catalysts."))),
        ("Principal risk", _compact_text(pm_sections.get("Things We're Watching", "Monitor breadth, financing conditions and the weakest constituent."))),
        ("Next session", _compact_text(pm_sections.get("Prepare for the Next Session", "Check the next company and UK macro calendar before the open."))),
    )
    executive_strip_html = "".join(
        '<td class="executive-cell" width="33.33%" valign="top" '
        'style="background:#102A47;padding:16px 17px;border-right:1px solid #2B4F73;">'
        '<div style="color:#46CFF5;font-size:9px;font-weight:800;letter-spacing:1.2px;'
        f'text-transform:uppercase;margin-bottom:6px;">{html.escape(label)}</div>'
        '<div style="color:#DCE8F4;font-size:12px;line-height:1.45;">'
        f'{html.escape(copy)}</div></td>'
        for label, copy in executive_items
    )
    pm_summary_html = "".join(
        ('<div style="margin:18px 0 10px;padding-top:14px;border-top:1px solid #31567E;'
         'color:#46CFF5;font-size:10px;font-weight:800;letter-spacing:1.4px;'
         'text-transform:uppercase;">Sector Summary</div>' if title == "Autos" else "")
        +
        '<div style="margin:0 0 16px;">'
        '<div style="margin:0 0 5px;color:#F6BD4A;font-size:10px;font-weight:800;'
        'letter-spacing:1px;text-transform:uppercase;">'
        f'{html.escape(title)}</div>'
        '<div style="color:#DCE8F4;font-size:14px;line-height:1.65;">'
        f'{html.escape(body)}</div></div>'
        for title, body in pm_summary
    )
    forward_links_html = "".join(
        f'<li style="margin:0 0 7px;"><a href="{html.escape(story["link"], quote=True)}" '
        f'style="color:#46CFF5;text-decoration:none;">{html.escape(story["title"])}</a>'
        f'<span style="color:#71869C;"> — {html.escape(story.get("source") or "publisher")}</span></li>'
        for story in forward_news
    )
    logo_data = base64.b64encode(LOGO_PATH.read_bytes()).decode("ascii")
    logo_src = f"data:image/jpeg;base64,{logo_data}"

    document = f"""<!doctype html>
<html><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<style>
@media only screen and (max-width:620px) {{
  .shell {{ width:100% !important; border-radius:0 !important; }}
  .content-pad {{ padding:18px 12px 8px !important; }}
  .header-pad {{ padding:20px 16px !important; }}
  .brand-logo-cell {{ width:82px !important; padding-right:14px !important; }}
  .brand-logo {{ width:74px !important; height:74px !important; border-radius:39px !important; }}
  .header-eyebrow {{ font-size:9px !important; letter-spacing:1.2px !important; }}
  .header-subtitle {{ font-size:12px !important; }}
  .executive-cell {{ display:block !important; width:100% !important; box-sizing:border-box !important; border-right:0 !important; border-bottom:1px solid #2B4F73 !important; }}
  .footer-pad {{ padding:18px 16px !important; }}
  .metric-cell {{ display:block !important; width:100% !important; box-sizing:border-box !important; }}
  .data-table {{ table-layout:fixed !important; }}
  .data-table th, .data-table td {{ padding:8px 4px !important; font-size:10px !important; overflow-wrap:anywhere !important; }}
  .headline {{ font-size:23px !important; }}
}}
</style></head>
<body style="margin:0;padding:0;background:#050D17;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Arial,sans-serif;color:#F7FAFC;">
<table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#050D17;"><tr><td align="center" style="padding:22px 10px;">
<table class="shell" role="presentation" width="100%" cellspacing="0" cellpadding="0" style="width:100%;max-width:760px;background:#091827;border:1px solid #1D3C5C;border-radius:20px;overflow:hidden;">
  <tr><td class="header-pad" style="background:#07121F;padding:24px 30px;border-bottom:1px solid #254A6B;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="table-layout:fixed;"><tr>
      <td class="brand-logo-cell" width="138" valign="middle" style="width:138px;padding-right:24px;"><img class="brand-logo" src="{logo_src}" alt="MelQuant Labs" width="106" height="106" style="display:block;width:106px;height:106px;border-radius:55px;border:2px solid #F6BD4A;"></td>
      <td valign="middle" style="min-width:0;overflow-wrap:break-word;">
        <div class="header-eyebrow" style="color:#F6BD4A;font-size:10px;font-weight:800;letter-spacing:2px;text-transform:uppercase;">MelQuant Labs · Closing Bell</div>
        <div class="headline" style="color:#F7FAFC;font-size:28px;font-weight:700;line-height:1.15;margin-top:6px;">{html.escape(report_title)}</div>
        <div class="header-subtitle" style="color:#9FB0C5;font-size:14px;margin-top:7px;">FTSE 100 &amp; Property Intelligence</div>
        <div style="color:#71869C;font-size:11px;margin-top:9px;">Data as of {html.escape(generated)}</div>
      </td>
    </tr></table>
  </td></tr>
  <tr><td style="padding:0;background:#2B4F73;border-bottom:1px solid #2B4F73;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>{executive_strip_html}</tr></table>
  </td></tr>
  <tr><td class="content-pad" style="padding:22px 25px 10px;">
    <div style="color:#8EA5BD;font-size:10px;font-weight:800;letter-spacing:1.4px;text-transform:uppercase;margin:0 5px 5px;">FTSE 100 Overview</div>
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>
      {_metric_card('FTSE 100', _format_pct(ftse_move), _move_colour(ftse_move), f'{advancers} up / {decliners} down')}
      {_metric_card('Market breadth', f'{advancers}/{decliners}', '#F7FAFC', 'advancers / decliners')}
      {_metric_card('Leader', _format_pct(leader_move), '#34D6A2', leader_name)}
      {_metric_card('Laggard', _format_pct(laggard_move), '#FF6B7A', laggard_name)}
    </tr></table>

    <div style="color:#46CFF5;font-size:10px;font-weight:800;letter-spacing:1.4px;text-transform:uppercase;margin:16px 5px 5px;">Property Developer Pulse</div>
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>
      {_metric_card('Property', _format_pct(real_estate_average), _move_colour(real_estate_average), 'listed bellwethers')}
      {_metric_card('Housebuilders', _format_pct(housebuilder_average), _move_colour(housebuilder_average), 'buyer-demand proxy')}
      {_metric_card('Materials', _format_pct(materials_average), _move_colour(materials_average), 'build-cost pipeline')}
      {_metric_card('SONIA', '—' if sonia is None else f'{sonia:.4f}%', '#F6BD4A', f'BoE · {sonia_as_of}')}
    </tr></table>

    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#102A47;border-left:4px solid #F6BD4A;border-radius:10px;margin:14px 5px 0;">
      <tr><td style="padding:15px 18px;color:#DCE8F4;font-size:14px;line-height:1.5;"><strong style="color:#FFFFFF;">Macro read:</strong> SONIA {'—' if sonia is None else f'{sonia:.4f}%'} · Bank Rate {'—' if bank_rate is None else f'{bank_rate:.2f}%'} · UK gilt proxy {_format_pct(gilt_move)} · GBP/USD {_format_pct(sterling_move)}. <strong style="color:#34D6A2;">FTSE leader:</strong> {html.escape(leader_name)} {_format_pct(leader_move)}. <strong style="color:#FF6B7A;">Laggard:</strong> {html.escape(laggard_name)} {_format_pct(laggard_move)}.</td></tr>
    </table>
    <div style="background:#0D243E;border:1px solid #274766;border-radius:9px;margin:10px 5px 18px;padding:10px 13px;color:#9FB0C5;font-size:10px;line-height:1.55;">
      <strong style="color:#F6BD4A;">FTSE 100 SECTOR WEIGHTS</strong> · {html.escape(weight_strip)}<br>
      <span style="color:#607990;">Proxy: iShares ISF holdings · {html.escape(sector_weights_as_of)}</span>
    </div>

    {_section_title('FTSE 100 tape', 'Leaders')}
    {_table(gainers, columns, emphasise_first=True)}
    {_section_title('FTSE 100 tape', 'Laggards')}
    {_table(fallers, columns, emphasise_first=True)}
    {_section_title('Catalyst radar', news_heading)}
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin:7px 0 18px;">{news_html}</table>
    {_section_title('Signal monitor', 'Notable FTSE 100 moves')}
    {_table(notable, columns)}
    {_section_title('Pillar 1 · Capital', 'Finance Cost & Credit Ticker')}
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin:10px 0 5px;"><tr>
      {_metric_card('SONIA', '—' if sonia is None else f'{sonia:.4f}%', '#F6BD4A', f'BoE · {sonia_as_of}')}
      {_metric_card('Bank Rate', '—' if bank_rate is None else f'{bank_rate:.2f}%', '#F7FAFC', f'BoE · {bank_rate_as_of}')}
      {_metric_card('UK gilts proxy', _format_pct(gilt_move), _move_colour(gilt_move), 'IGLT daily move')}
      {_metric_card('GBP/USD', _format_pct(sterling_move), _move_colour(sterling_move), 'currency pressure')}
    </tr></table>
    {_table(banks.sort_values('day_pct', ascending=False), columns, emphasise_extremes=True)}
    {_section_title('Pillar 2 · Supply chain', 'Material Inflation Pipeline')}
    {_table(materials.sort_values('day_pct', ascending=False), columns, emphasise_extremes=True)}
    {_section_title('Pillar 3 · Asset values', 'Real Estate Bellwethers')}
    {_table(real_estate.sort_values('day_pct', ascending=False), columns, emphasise_extremes=True)}
    {_section_title('Pillar 3 · Competition', 'Housebuilders & Buyer Demand')}
    {_table(housebuilders.sort_values('day_pct', ascending=False), columns, emphasise_extremes=True)}
    {_section_title('Pillar 4 · Affordability', 'Consumer Health & Spending Power')}
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="margin:10px 0 5px;"><tr>
      {_metric_card('Discretionary', _format_pct(discretionary_average), _move_colour(discretionary_average), 'Next · M&S · Kingfisher')}
      {_metric_card('Staples', _format_pct(staples_average), _move_colour(staples_average), 'Tesco · Sainsbury · Unilever')}
      {_metric_card('FTSE breadth', f'{advancers}/{decliners}', '#F7FAFC', 'advancers / decliners')}
      {_metric_card('Notable moves', str(len(notable)), '#F6BD4A', 'price or volume flags')}
    </tr></table>
    {_table(consumers.sort_values('day_pct', ascending=False), columns, emphasise_extremes=True)}
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#102A47;border-left:4px solid #46CFF5;border-radius:10px;margin:10px 5px 16px;">
      <tr><td style="padding:15px 18px;color:#DCE8F4;font-size:13px;line-height:1.55;"><strong style="color:#FFFFFF;">Lotus Noor Developments Focus Lens:</strong> {html.escape(property_read)}</td></tr>
    </table>
    {_section_title('Research hypotheses', 'Alpha View')}
    {alpha_view_html}
    <div style="color:#71869C;font-size:10px;line-height:1.5;margin:8px 0 22px;">Research hypotheses for monitoring only—not recommendations or personal investment advice. Verify market data, news and catalyst timing at source.</div>
    {_section_title('Cross-market context', 'Drivers dashboard')}
    {_table(drivers, columns)}
    {_section_title('Portfolio manager close', 'Daily PM Summary')}
    <table id="daily-pm-summary" role="presentation" width="100%" cellspacing="0" cellpadding="0" style="background:#102A47;border:1px solid #31567E;border-left:4px solid #F6BD4A;border-radius:12px;margin:10px 0 10px;">
      <tr><td style="padding:18px 20px;">{pm_summary_html}</td></tr>
    </table>
    {f'<div style="color:#8EA5BD;font-size:10px;font-weight:800;letter-spacing:1px;text-transform:uppercase;margin:14px 0 7px;">Forward-watch sources</div><ul style="margin:0 0 25px;padding-left:20px;color:#9FB0C5;font-size:11px;line-height:1.5;">{forward_links_html}</ul>' if forward_links_html else ''}
  </td></tr>
  <tr><td class="footer-pad" style="background:#07121F;padding:22px 30px;border-top:1px solid #254A6B;">
    <table role="presentation" width="100%" cellspacing="0" cellpadding="0"><tr>
      <td valign="middle" style="color:#F6BD4A;font-size:12px;font-weight:800;letter-spacing:1px;">MELQUANT LABS</td>
      <td align="right" style="color:#71869C;font-size:10px;line-height:1.5;">Prices: Yahoo Finance adjusted daily data · market date {html.escape(market_as_of)}<br>Rates: Bank of England · weights: iShares ISF ({html.escape(sector_weights_as_of)})<br>News: Google News RSS headlines only—not article bodies. Verify every claim at source.</td>
    </tr></table>
  </td></tr>
</table></td></tr></table></body></html>"""
    frames = {
        "FTSE 100": ftse,
        "Capital & Finance": banks,
        "Autos Summary": autos,
        "Technology Summary": technology,
        "Real Estate": real_estate,
        "Housebuilders": housebuilders,
        "Materials & Infrastructure": materials,
        "Consumer Health": consumers,
        "Market Drivers": drivers,
        "Sector Moves": sector_moves,
        "Property & Infrastructure": property_names,
    }
    return document, frames, news


def write_workbook(frames: dict[str, pd.DataFrame], news: list[dict]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, frame in frames.items():
        sheet = workbook.create_sheet(name[:31])
        output = frame.reset_index(names="ticker")
        sheet.append(list(output.columns))
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="17365D")
        for row in output.itertuples(index=False, name=None):
            sheet.append(list(row))
    sheet = workbook.create_sheet("News")
    sheet.append(["Headline", "Link", "Published"])
    for story in news:
        published = story["published"].isoformat() if story.get("published") else ""
        sheet.append([story["title"], story["link"], published])
    workbook.save(WORKBOOK_OUTPUT)


def validate_delivery(frames: dict[str, pd.DataFrame], now: datetime | None = None) -> str:
    """Refuse to email a report whose core FTSE tape is incomplete or stale."""
    current = (now or datetime.now().astimezone()).astimezone()
    ftse = frames.get("FTSE 100", pd.DataFrame())
    if len(ftse) < 90:
        raise ValueError(f"FTSE coverage is incomplete: received {len(ftse)} constituents")
    dates = pd.to_datetime(ftse.get("as_of"), errors="coerce").dropna()
    if dates.empty:
        raise ValueError("FTSE market date is unavailable")
    market_date = dates.max().date()
    if market_date != current.date():
        raise ValueError(
            f"FTSE data is stale: latest market date {market_date.isoformat()}, "
            f"expected {current.date().isoformat()}"
        )
    coverage = int((dates.dt.date == market_date).sum())
    if coverage < 90:
        raise ValueError(
            f"FTSE same-day coverage is incomplete: {coverage} constituents for {market_date}"
        )
    return market_date.isoformat()


def validate_stable_close(
    frames: dict[str, pd.DataFrame],
    now: datetime | None = None,
    state_path: Path = CLOSE_CHECK_PATH,
    minimum_minutes: int = 3,
) -> str:
    """Require two identical post-close FTSE snapshots several minutes apart."""
    current = (now or datetime.now().astimezone()).astimezone()
    if current.time().replace(tzinfo=None) < time(16, 50):
        raise ValueError("market-close delivery is blocked before 16:50 London time")
    market_date = validate_delivery(frames, current)
    ftse = frames["FTSE 100"].sort_index()
    payload = [
        (str(ticker), str(row["as_of"]), round(float(row["price"]), 6))
        for ticker, row in ftse.iterrows()
    ]
    signature = hashlib.sha256(
        json.dumps(payload, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    previous = {}
    try:
        previous = json.loads(state_path.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        pass
    checked_at = datetime.fromisoformat(previous["checked_at"]) if previous.get("checked_at") else None
    stable = (
        previous.get("market_date") == market_date
        and previous.get("signature") == signature
        and checked_at is not None
        and (current - checked_at).total_seconds() >= minimum_minutes * 60
    )
    first_seen = (
        previous["checked_at"]
        if previous.get("market_date") == market_date
        and previous.get("signature") == signature
        and previous.get("checked_at")
        else current.isoformat()
    )
    state_path.write_text(
        json.dumps(
            {"market_date": market_date, "signature": signature, "checked_at": first_seen},
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )
    if not stable:
        raise ValueError(
            f"waiting for a second unchanged FTSE snapshot at least {minimum_minutes} minutes apart"
        )
    return market_date


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Build files without emailing")
    args = parser.parse_args()

    document, frames, news = build_digest()
    HTML_OUTPUT.write_text(document, encoding="utf-8")
    write_workbook(frames, news)
    print(f"[ok] HTML briefing written: {HTML_OUTPUT}")
    print(f"[ok] Excel dashboard written: {WORKBOOK_OUTPUT}")
    if not args.dry_run:
        market_date = validate_stable_close(frames)
        print(f"[ok] stable-close validation passed: FTSE market date {market_date}")
        subject_date = datetime.now().strftime("%d/%m/%Y")
        email_document = document.replace(
            f"data:image/jpeg;base64,{base64.b64encode(LOGO_PATH.read_bytes()).decode('ascii')}",
            "cid:melquantlabs-logo",
        )
        send_email(
            f"MelQuant Labs Daily Market Briefing | {subject_date}",
            email_document,
            inline_images={"melquantlabs-logo": LOGO_PATH},
        )


if __name__ == "__main__":
    main()
