import unittest
import os
from email import message_from_string
from unittest.mock import patch

from openpyxl import Workbook

import market_monitor


class EvaluateQuotesTests(unittest.TestCase):
    def test_first_large_move_triggers_once(self):
        quotes = {
            "AUTO.L": {"price": 103.0, "prev_close": 100.0, "day_pct": 3.0}
        }
        state = {}

        alerts, moves = market_monitor.evaluate_quotes(quotes, state, 3.0)

        self.assertEqual(len(alerts), 1)
        self.assertAlmostEqual(moves["AUTO.L"], 3.0)
        self.assertTrue(state["AUTO.L"]["alert_active"])
        self.assertEqual(state["AUTO.L"]["last_price"], 103.0)

        repeated_alerts, _ = market_monitor.evaluate_quotes(quotes, state, 3.0)
        self.assertEqual(repeated_alerts, [])

    def test_alert_resets_after_move_returns_below_threshold(self):
        state = {"AUTO.L": {"last_price": 103.0, "alert_active": True}}
        quiet_quote = {
            "AUTO.L": {"price": 103.1, "prev_close": 103.0, "day_pct": 0.1}
        }

        market_monitor.evaluate_quotes(quiet_quote, state, 3.0)

        self.assertFalse(state["AUTO.L"]["alert_active"])

    def test_dashboard_uses_pre_update_move(self):
        workbook = Workbook()
        workbook.remove(workbook.active)
        quotes = {
            "AUTO.L": {"price": 103.0, "prev_close": 100.0, "day_pct": 3.0}
        }
        state = {"AUTO.L": {"last_price": 103.0, "alert_active": True}}

        market_monitor.update_live_sheet(
            workbook,
            quotes,
            state,
            {"AUTO.L"},
            {"AUTO.L": 3.0},
        )

        sheet = workbook["Live"]
        self.assertEqual(sheet.cell(row=2, column=4).value, 3.0)
        self.assertEqual(sheet.cell(row=2, column=5).value, "ALERT")


class EmailPrivacyTests(unittest.TestCase):
    @patch("market_monitor.smtplib.SMTP")
    def test_sender_brand_and_recipient_list_are_private(self, smtp):
        settings = {
            "SMTP_HOST": "smtp.example.com",
            "SMTP_PORT": "587",
            "SMTP_SECURITY": "starttls",
            "SMTP_USER": "sender@example.com",
            "SMTP_PASSWORD": "private-password",
            "EMAIL_FROM": "sender@example.com",
            "EMAIL_FROM_NAME": "MelQuant Labs",
            "EMAIL_TO": "first@example.com,second@example.com",
        }
        with patch.dict(os.environ, settings, clear=False):
            market_monitor.send_email("Daily briefing", "<p>Report</p>")

        server = smtp.return_value
        envelope_from, envelope_to, raw_message = server.sendmail.call_args.args
        message = message_from_string(raw_message)

        self.assertEqual(envelope_from, "sender@example.com")
        self.assertEqual(envelope_to, ["first@example.com", "second@example.com"])
        self.assertEqual(message["From"], "MelQuant Labs <sender@example.com>")
        self.assertEqual(message["To"], "Undisclosed recipients:;")
        self.assertIsNone(message["Bcc"])


if __name__ == "__main__":
    unittest.main()
