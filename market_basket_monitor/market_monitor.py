#!/usr/bin/env python3
"""
Market basket monitor.

Two modes:
  --mode check    Fast price check. Fires macOS notification + email if any
                   ticker has moved >= ALERT_THRESHOLD_PCT since the last run
                   (meant to be run frequently, e.g. every 15-30 min during market hours).
  --mode weekly   Builds and emails a weekly summary: price move over the last
                   7 days for each ticker, plus recent news headlines.

Run manually to test:
    python3 market_monitor.py --mode check
    python3 market_monitor.py --mode weekly
"""
import argparse
import html
import json
import os
import smtplib
import subprocess
import sys
from datetime import datetime, timedelta, timezone
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import yfinance as yf
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

import config

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

STATE_PATH = BASE_DIR / config.STATE_FILE
DASHBOARD_PATH = BASE_DIR / config.DASHBOARD_FILE


# ---------- state ----------

def load_state():
    if STATE_PATH.exists():
        try:
            return json.loads(STATE_PATH.read_text())
        except json.JSONDecodeError:
            return {}
    return {}


def save_state(state):
    STATE_PATH.write_text(json.dumps(state, indent=2))


# ---------- prices ----------

def _fast_info_value(fast, *keys):
    """yfinance's fast_info supports both attribute and dict-style access depending on
    version, and key naming has changed across releases. Try every variant we know of."""
    for key in keys:
        try:
            val = fast[key]
            if val is not None:
                return val
        except Exception:
            pass
        val = getattr(fast, key, None)
        if val is not None:
            return val
    return None


def fetch_quotes(tickers):
    """Return {ticker: {"price": float, "prev_close": float, "day_pct": float}}"""
    quotes = {}
    for t in tickers:
        try:
            tk = yf.Ticker(t)
            fast = tk.fast_info
            price = _fast_info_value(fast, "last_price", "lastPrice", "regularMarketPrice")
            prev_close = _fast_info_value(fast, "previous_close", "previousClose", "regularMarketPreviousClose")

            # Fall back to recent daily history if fast_info didn't have what we need.
            if price is None or prev_close is None:
                hist = tk.history(period="5d")
                if not hist.empty:
                    price = price if price is not None else float(hist["Close"].iloc[-1])
                    prev_close = prev_close if prev_close is not None else float(hist["Close"].iloc[-2])

            if price is None or prev_close is None:
                raise ValueError("missing price data")
            day_pct = (price - prev_close) / prev_close * 100
            quotes[t] = {"price": float(price), "prev_close": float(prev_close), "day_pct": day_pct}
        except Exception as e:
            print(f"[warn] could not fetch {t}: {e}", file=sys.stderr)
    return quotes


def fetch_week_history(ticker):
    """Return (start_price, end_price, pct_change) over the last 7 calendar days."""
    tk = yf.Ticker(ticker)
    hist = tk.history(period="8d")
    if hist.empty:
        return None
    start_price = float(hist["Close"].iloc[0])
    end_price = float(hist["Close"].iloc[-1])
    pct = (end_price - start_price) / start_price * 100
    return start_price, end_price, pct


# ---------- news ----------

def fetch_news(query, max_items=3, days_back=None):
    """Pull headlines from Google News RSS for a query. No API key required."""
    import feedparser
    import requests
    import urllib.parse

    url = f"https://news.google.com/rss/search?q={urllib.parse.quote(query)}&hl=en-GB&gl=GB&ceid=GB:en"
    response = requests.get(
        url,
        timeout=20,
        headers={"User-Agent": "MelquantLabsMarketMonitor/1.0"},
    )
    response.raise_for_status()
    feed = feedparser.parse(response.content)
    items = []
    cutoff = None
    if days_back:
        cutoff = datetime.now(timezone.utc) - timedelta(days=days_back)
    for entry in feed.entries:
        published = None
        if getattr(entry, "published_parsed", None):
            published = datetime(*entry.published_parsed[:6], tzinfo=timezone.utc)
        if cutoff and published and published < cutoff:
            continue
        source = ""
        if getattr(entry, "source", None):
            source = str(getattr(entry.source, "title", "") or "").strip()
        items.append(
            {
                "title": entry.title,
                "link": entry.link,
                "published": published,
                "source": source,
            }
        )
        if len(items) >= max_items:
            break
    return items


# ---------- delivery ----------

def send_macos_notification(title, message):
    if not config.MACOS_NOTIFICATIONS:
        return
    try:
        script = f'display notification "{message}" with title "{title}"'
        subprocess.run(["osascript", "-e", script], check=False)
    except FileNotFoundError:
        pass  # not on macOS


def send_email(subject, html_body, inline_images=None):
    host = os.environ.get("SMTP_HOST")
    port = int(os.environ.get("SMTP_PORT", "587"))
    security = os.environ.get("SMTP_SECURITY", "starttls").lower()
    user = os.environ.get("SMTP_USER") or os.environ.get("BT_EMAIL_USER")
    password = os.environ.get("SMTP_PASSWORD") or os.environ.get("BT_EMAIL_PASSWORD")
    email_from = os.environ.get("EMAIL_FROM", user)
    email_to = os.environ.get("EMAIL_TO", user)
    if not host or not user or not password or not email_from or not email_to:
        print("[error] Email settings are incomplete (check your .env file). "
              "Skipping email send.", file=sys.stderr)
        return
    if inline_images:
        msg = MIMEMultipart("related")
        alternative = MIMEMultipart("alternative")
        alternative.attach(MIMEText("This briefing requires an HTML-capable email client.", "plain"))
        alternative.attach(MIMEText(html_body, "html"))
        msg.attach(alternative)
        for content_id, image_path in inline_images.items():
            with open(image_path, "rb") as image_file:
                image_part = MIMEImage(image_file.read())
            image_part.add_header("Content-ID", f"<{content_id}>")
            image_part.add_header("Content-Disposition", "inline", filename=Path(image_path).name)
            msg.attach(image_part)
    else:
        msg = MIMEText(html_body, "html")
    msg["Subject"] = subject
    msg["From"] = email_from
    msg["To"] = email_to

    recipients = [address.strip() for address in email_to.split(",") if address.strip()]
    if security == "ssl":
        server = smtplib.SMTP_SSL(host, port)
    elif security == "starttls":
        server = smtplib.SMTP(host, port)
        server.ehlo()
        server.starttls()
        server.ehlo()
    else:
        raise ValueError("SMTP_SECURITY must be 'ssl' or 'starttls'")

    with server:
        server.login(user, password)
        server.sendmail(email_from, recipients, msg.as_string())
    print(f"[ok] email sent: {subject}")


# ---------- excel dashboard ----------

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")


def _open_workbook():
    if DASHBOARD_PATH.exists():
        try:
            return load_workbook(DASHBOARD_PATH)
        except Exception as e:
            print(f"[warn] could not open existing dashboard ({e}); creating a new one", file=sys.stderr)
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def _ensure_sheet(wb, name, headers):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    return ws


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 60)


def _clear_data_rows(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def update_live_sheet(wb, quotes, state, alerted_tickers, moves_since_last=None):
    ws = _ensure_sheet(wb, "Live", ["Ticker", "Price", "Day %", "Since Last Check %", "Alert", "Updated"])
    _clear_data_rows(ws)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    moves_since_last = moves_since_last or {}
    for ticker, q in quotes.items():
        reference_price = state.get(ticker, {}).get("last_price", q["prev_close"])
        move_since_last = moves_since_last.get(
            ticker,
            (q["price"] - reference_price) / reference_price * 100,
        )
        ws.append([
            ticker, round(q["price"], 2), round(q["day_pct"], 2), round(move_since_last, 2),
            "ALERT" if ticker in alerted_tickers else "", now,
        ])
        r = ws.max_row
        ws.cell(row=r, column=3).fill = GREEN if q["day_pct"] >= 0 else RED
        if ticker in alerted_tickers:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = YELLOW
    _autosize(ws)


def append_weekly_history(wb, rows):
    ws = _ensure_sheet(wb, "Weekly History", ["Week Ending", "Ticker", "Start Price", "End Price", "Change %"])
    week_ending = datetime.now().strftime("%Y-%m-%d")
    for ticker, start, end, pct, _ in rows:
        ws.append([week_ending, ticker, round(start, 2), round(end, 2), round(pct, 2)])
        r = ws.max_row
        ws.cell(row=r, column=5).fill = GREEN if pct >= 0 else RED
    _autosize(ws)


def upsert_news(wb, news_by_ticker):
    ws = _ensure_sheet(wb, "News", ["Ticker", "Headline", "Link", "Published"])
    for ticker in news_by_ticker:
        for r in reversed(range(2, ws.max_row + 1)):
            if ws.cell(row=r, column=1).value == ticker:
                ws.delete_rows(r)
    for ticker, items in news_by_ticker.items():
        for n in items:
            published = n["published"].strftime("%Y-%m-%d %H:%M") if n["published"] else ""
            ws.append([ticker, n["title"], n["link"], published])
    _autosize(ws)


def save_workbook(wb):
    try:
        wb.save(DASHBOARD_PATH)
        print(f"[ok] dashboard updated: {DASHBOARD_PATH}")
    except PermissionError:
        print(f"[warn] {DASHBOARD_PATH} is open/locked (e.g. open in Excel) — "
              f"close it in Excel so the next run can save.", file=sys.stderr)


# ---------- modes ----------

def evaluate_quotes(quotes, state, threshold_pct):
    """Update monitoring state and return newly triggered alerts and price moves."""
    alerts = []
    moves_since_last = {}

    for ticker, quote in quotes.items():
        ticker_state = state.setdefault(ticker, {})
        reference_price = ticker_state.get("last_price", quote["prev_close"])
        move_since_last = (
            (quote["price"] - reference_price) / reference_price * 100
        )
        moves_since_last[ticker] = move_since_last

        threshold_reached = (
            abs(quote["day_pct"]) >= threshold_pct
            or abs(move_since_last) >= threshold_pct
        )
        if threshold_reached and not ticker_state.get("alert_active", False):
            alerts.append((ticker, quote, move_since_last))

        ticker_state["alert_active"] = threshold_reached
        ticker_state["last_price"] = quote["price"]

    return alerts, moves_since_last

def run_check():
    state = load_state()
    quotes = fetch_quotes(config.TICKERS)
    alerts, moves_since_last = evaluate_quotes(
        quotes,
        state,
        config.ALERT_THRESHOLD_PCT,
    )

    save_state(state)

    # Always refresh the Excel dashboard's Live sheet, alerts or not.
    wb = _open_workbook()
    alerted_tickers = {t for t, _, _ in alerts}
    update_live_sheet(
        wb,
        quotes,
        state,
        alerted_tickers,
        moves_since_last,
    )

    if not alerts:
        save_workbook(wb)
        print("[ok] no alerts this run")
        return

    lines = []
    news_by_ticker = {}
    for ticker, q, move_since_last in alerts:
        direction = "up" if q["day_pct"] >= 0 else "down"
        lines.append(
            f"{ticker}: {q['price']:.2f} ({direction} {q['day_pct']:+.2f}% today, "
            f"{move_since_last:+.2f}% since last check)"
        )
        news = fetch_news(ticker, max_items=config.NEWS_ITEMS_PER_TICKER)
        news_by_ticker[ticker] = news
        for n in news:
            lines.append(f"    - {n['title']} ({n['link']})")

    upsert_news(wb, news_by_ticker)
    save_workbook(wb)

    summary = "\n".join(lines)
    print("[alert]\n" + summary)

    send_macos_notification(
        f"{config.BASKET_NAME}: {len(alerts)} alert(s)",
        "; ".join(f"{t}: {q['day_pct']:+.2f}%" for t, q, _ in alerts),
    )

    html_body = "<h3>Basket alert</h3><pre>" + html.escape(summary) + "</pre>"
    send_email(
        f"[Market Alert] {len(alerts)} move(s) in {config.BASKET_NAME}",
        html_body,
    )


def run_weekly():
    rows = []
    news_by_ticker = {}
    for ticker in config.TICKERS:
        hist = fetch_week_history(ticker)
        if hist is None:
            continue
        start, end, pct = hist
        news = fetch_news(ticker, max_items=config.NEWS_ITEMS_PER_TICKER,
                           days_back=config.WEEKLY_NEWS_LOOKBACK_DAYS)
        rows.append((ticker, start, end, pct, news))
        news_by_ticker[ticker] = news

    html_parts = [f"<h2>Weekly basket summary — {config.BASKET_NAME}</h2>",
                  f"<p>{datetime.now().strftime('%A %d %B %Y')}</p>",
                  "<table border='1' cellpadding='6' cellspacing='0'>",
                  "<tr><th>Ticker</th><th>7d ago</th><th>Now</th><th>Change</th></tr>"]
    for ticker, start, end, pct, _ in rows:
        color = "green" if pct >= 0 else "red"
        html_parts.append(
            f"<tr><td>{ticker}</td><td>{start:.2f}</td><td>{end:.2f}</td>"
            f"<td style='color:{color}'>{pct:+.2f}%</td></tr>"
        )
    html_parts.append("</table>")

    html_parts.append("<h3>Recent news</h3>")
    for ticker, _, _, _, news in rows:
        html_parts.append(f"<p><b>{html.escape(ticker)}</b></p><ul>")
        if not news:
            html_parts.append("<li>No notable headlines this week.</li>")
        for n in news:
            html_parts.append(
                f'<li><a href="{html.escape(n["link"], quote=True)}">'
                f'{html.escape(n["title"])}</a></li>'
            )
        html_parts.append("</ul>")

    html = "\n".join(html_parts)
    print("[ok] built weekly digest")
    send_email(f"Weekly basket summary — {config.BASKET_NAME}", html)

    # Refresh the whole dashboard: current prices, weekly trend row, and news.
    state = load_state()
    quotes = fetch_quotes(config.TICKERS)
    wb = _open_workbook()
    update_live_sheet(wb, quotes, state, alerted_tickers=set())
    append_weekly_history(wb, rows)
    upsert_news(wb, news_by_ticker)
    save_workbook(wb)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["check", "weekly"], required=True)
    args = parser.parse_args()

    if args.mode == "check":
        run_check()
    else:
        run_weekly()


if __name__ == "__main__":
    main()
